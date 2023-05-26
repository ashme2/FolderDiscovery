# -*- coding: utf-8 -*-

# Form implementation generated from reading ui file 'FolDis-MainWindow_V0.03.ui'
#
# Created by: PyQt5 UI code generator 5.13.0
#
# WARNING! All changes made in this file will be lost!


from PyQt5 import QtCore, QtGui, QtWidgets
from PyQt5.QtWidgets import QFileDialog
import os, sys
import pandas as pd
from PyQt5.QtCore import QAbstractTableModel, Qt
from PyQt5.QtGui import QIcon, QPixmap
from PyQt5.QtWidgets import QApplication, QWidget, QLabel
import ctypes

from openpyxl import Workbook, load_workbook
from PyQt5.QtWidgets import QPushButton, QMessageBox
from PyQt5.QtCore import pyqtSlot


folPath = ""
Version = 'V1.0'

# Program Version Data
class pandasModel(QAbstractTableModel):

    def __init__(self, data):
        QAbstractTableModel.__init__(self)
        self._data = data

    def rowCount(self, parent=None):
        return self._data.shape[0]

    def columnCount(self, parnet=None):
        return self._data.shape[1]

    def data(self, index, role=Qt.DisplayRole):
        if index.isValid():
            if role == Qt.DisplayRole:
                return str(self._data.iloc[index.row(), index.column()])
        return None

    def headerData(self, col, orientation, role):
        if orientation == Qt.Horizontal and role == Qt.DisplayRole:
            return self._data.columns[col]
        return None

# Continue Question
def showDialog(msg, title):
    msgBox = QMessageBox()
    msgBox.setIcon(QMessageBox.Question)
    msgBox.setText(msg)
    msgBox.setWindowTitle(title)
    msgBox.setStandardButtons(QMessageBox.Ok | QMessageBox.Cancel)
    contBtn = msgBox.button(QMessageBox.Ok)
    contBtn.setText('Continue')
    stpBtn = msgBox.button(QMessageBox.Cancel)
    stpBtn.setText('Stop')
    
    # Create Widget Icon Image 
    exePath = os.getcwd()
    iconPath = os.path.join(exePath, "System-icon.png")
    icon = QtGui.QIcon()
    icon.addPixmap(QtGui.QPixmap(iconPath), QtGui.QIcon.Normal, QtGui.QIcon.Off)
    msgBox.setWindowIcon(icon)
    myappid = 'mycompany.myproduct.subproduct.version' # arbitrary string
    ctypes.windll.shell32.SetCurrentProcessExplicitAppUserModelID(myappid)
    
    returnValue = msgBox.exec()
    return returnValue

class Ui_MainWindow(object):
    def setupUi(self, MainWindow):
        MainWindow.setObjectName("MainWindow")
        MainWindow.resize(502, 410)
        self.centralwidget = QtWidgets.QWidget(MainWindow)
        self.centralwidget.setObjectName("centralwidget")
        self.selFolBtn = QtWidgets.QPushButton(self.centralwidget)
        self.selFolBtn.setGeometry(QtCore.QRect(20, 20, 101, 41))
        self.selFolBtn.setObjectName("selFolBtn")
        self.label_2 = QtWidgets.QLabel(self.centralwidget)
        self.label_2.setGeometry(QtCore.QRect(20, 70, 101, 31))
        self.label_2.setLayoutDirection(QtCore.Qt.LeftToRight)
        self.label_2.setAlignment(QtCore.Qt.AlignCenter)
        self.label_2.setObjectName("label_2")
        self.ExtLineEdit = QtWidgets.QLineEdit(self.centralwidget)
        self.ExtLineEdit.setGeometry(QtCore.QRect(140, 70, 161, 31))
        self.ExtLineEdit.setObjectName("ExtLineEdit")
        self.label_3 = QtWidgets.QLabel(self.centralwidget)
        self.label_3.setGeometry(QtCore.QRect(20, 110, 260, 20))
        self.label_3.setAlignment(QtCore.Qt.AlignCenter)
        self.label_3.setObjectName("label_3")
        self.folDisBtn = QtWidgets.QPushButton(self.centralwidget)
        self.folDisBtn.setGeometry(QtCore.QRect(320, 70, 161, 31))
        self.folDisBtn.setObjectName("folDisBtn")
        self.textBrowser = QtWidgets.QTextBrowser(self.centralwidget)
        self.textBrowser.setGeometry(QtCore.QRect(140, 20, 341, 41))
        self.textBrowser.setObjectName("textBrowser")
        self.imageLbl = QtWidgets.QLabel(self.centralwidget)
        self.imageLbl.setGeometry(QtCore.QRect(320, 255, 161, 126))
        self.imageLbl.setObjectName("imageLbl")
        self.tableView = QtWidgets.QTableView(self.centralwidget)
        self.tableView.setGeometry(QtCore.QRect(20, 210, 292, 175))
        font = QtGui.QFont()
        font.setPointSize(8)
        self.tableView.setFont(font)
        self.tableView.setObjectName("tableView")
        self.radioButton = QtWidgets.QRadioButton(self.centralwidget)
        self.radioButton.setGeometry(QtCore.QRect(335, 210, 95, 20))
        self.radioButton.setObjectName("radioButton")
        self.radioButton_2 = QtWidgets.QRadioButton(self.centralwidget)
        self.radioButton_2.setGeometry(QtCore.QRect(335, 230, 95, 20))
        self.radioButton_2.setObjectName("radioButton_2")
        MainWindow.setCentralWidget(self.centralwidget)
        self.menubar = QtWidgets.QMenuBar(MainWindow)
        self.menubar.setGeometry(QtCore.QRect(0, 0, 502, 26))
        self.menubar.setObjectName("menubar")
        MainWindow.setMenuBar(self.menubar)
        self.statusbar = QtWidgets.QStatusBar(MainWindow)
        self.statusbar.setObjectName("statusbar")
        MainWindow.setStatusBar(self.statusbar)

        self.retranslateUi(MainWindow)
        QtCore.QMetaObject.connectSlotsByName(MainWindow)       
        
        self.radioButton.setChecked(True)

        # Edit MainWindow
        self.selFolBtn.clicked.connect(self.selFol)
        self.folDisBtn.clicked.connect(self.folDis)
        exePath = os.getcwd()
        logoPath = os.path.join(exePath, "EagleX-2.1-FolDis.png")
        if logoPath:
            pixmap = QtGui.QPixmap(logoPath)
            pixmap = pixmap.scaled(self.imageLbl.width(), self.imageLbl.height(), QtCore.Qt.KeepAspectRatio)
            self.imageLbl.setPixmap(pixmap)
            self.imageLbl.setAlignment(QtCore.Qt.AlignCenter)
        # Program Version Data
        progData = pd.DataFrame({'Program': ['Program Name', 'Version', 'Date', 'Author', 'Email'],
                    'Data': ['Folder Discovery', Version, '19/05/2023', 'Eng. Ashraf Ibrahim', 'eaglexengineering@gmail.com']})
        model = pandasModel(progData)
        self.tableView.setModel(model)
        self.tableView.resizeColumnsToContents()
        self.tableView.resizeRowsToContents()

    def retranslateUi(self, MainWindow):
        _translate = QtCore.QCoreApplication.translate
        MainWindow.setWindowTitle(_translate("MainWindow", "Folder Discovery"))
        self.selFolBtn.setText(_translate("MainWindow", "Select Folder"))
        self.label_2.setText(_translate("MainWindow", "Write extension"))
        self.label_3.setText(_translate("MainWindow", "to search for, like: (.dwg - .exe - .xlsx - .jpg)"))
        self.folDisBtn.setText(_translate("MainWindow", "Folder Discovery"))
        self.imageLbl.setText(_translate("MainWindow", "TextLabel"))
        self.radioButton.setText(_translate("MainWindow", "Text Output"))
        self.radioButton_2.setText(_translate("MainWindow", "Excel Output"))
        
        # Create Widget Icon Image 
        exePath = os.getcwd()
        iconPath = os.path.join(exePath, "Search-icon.png")
        icon = QtGui.QIcon()
        icon.addPixmap(QtGui.QPixmap(iconPath), QtGui.QIcon.Normal, QtGui.QIcon.Off)
        MainWindow.setWindowIcon(icon)
        myappid = 'mycompany.myproduct.subproduct.version' # arbitrary string
        ctypes.windll.shell32.SetCurrentProcessExplicitAppUserModelID(myappid)
        
        
    # Define Functions        
    def selFol(self):
        fileName = QFileDialog.getExistingDirectory(None, "Select Folder", "", QtWidgets.QFileDialog.ShowDirsOnly)
        if fileName:
            self.textBrowser.setText(fileName.replace('/', '\\'))
            global folPath
            folPath = fileName.replace('/', '\\')
    
    def folDis(self):
        # This Function Write The File Path For Selected Folder
        self.selFolBtn.setEnabled(False)
        self.folDisBtn.setEnabled(False)
        self.radioButton.setEnabled(False)
        self.radioButton_2.setEnabled(False)
        self.ExtLineEdit.setEnabled(False)
        self.textBrowser.setEnabled(False)
        ##############################

        st = []
        stFilePath = []
        stFolderPath = []
        stFileName =[]
        i = 0
        filePath = folPath
        k = 0
        qsMsg = 0
        for folderName, subfolders, filenames in os.walk(filePath):
            for filename in filenames:
                if ((k % 100000) == 0) and (k != 0):
                    title = 'Program Overload'
                    msg = 'You reach file no. {}.\nDo you want to continue, or stop?'.format(k)
                    qsMsg = showDialog(msg, title)
                    if qsMsg == 4194304:
                        break
                extTxt = self.ExtLineEdit.text()
                if extTxt != "":
                    extLen = len(extTxt) * -1
                    if filename[extLen :] == extTxt:
                        st.append (folderName + ': ' + filename)
                        stFilePath.append (folderName + '\\' + filename)
                        stFolderPath.append (folderName)
                        stFileName.append (filename)
                else:
                    st.append (folderName + ': ' + filename)
                    stFilePath.append (folderName + '\\' + filename)
                    stFolderPath.append (folderName)
                    stFileName.append (filename)
                k += 1
            if qsMsg == 4194304:
                break
        if self.radioButton.isChecked():
            file = open("Folder-Discovery.txt","w+")
            for x in st:
                file.write(x + "\n")
            file.close()
        else:
            wb=Workbook()
            xlfilepath = "Folder-Discovery.xlsx"
            wb.save(xlfilepath)
            wb = load_workbook(xlfilepath)
            wb.create_sheet('Path')
            wb.create_sheet('Hyperlink')
            wb.remove(wb.get_sheet_by_name('Sheet'))
            sheet1 = wb.get_sheet_by_name('Path')
            sheet2 = wb.get_sheet_by_name('Hyperlink')
            sheet1.cell(row=1, column=1).value = 'File Path'
            sheet1.cell(row=1, column=2).value = 'Folder Path'
            sheet1.cell(row=1, column=3).value = 'File Name'
            for i in range(len(st)):
                sheet1.cell(row=i+2, column=1).value = stFilePath[i]
                sheet1.cell(row=i+2, column=2).value = stFolderPath[i]
                sheet1.cell(row=i+2, column=3).value = stFileName[i]
            sheet2.cell(row=1, column=1).value = 'File Path'
            sheet2.cell(row=1, column=2).value = 'Folder Path'
            sheet2.cell(row=1, column=3).value = 'File Name'
            for i in range(len(st)):
                sheet2.cell(row=i+2, column=1).value = '=HYPERLINK("{}")'.format(stFilePath[i])
                sheet2.cell(row=i+2, column=2).value = '=HYPERLINK("{}")'.format(stFolderPath[i])
                sheet2.cell(row=i+2, column=3).value = stFileName[i]
            wb.save(xlfilepath)
        self.selFolBtn.setEnabled(True)
        self.folDisBtn.setEnabled(True)
        self.radioButton.setEnabled(True)
        self.radioButton_2.setEnabled(True)
        self.ExtLineEdit.setEnabled(True)
        self.textBrowser.setEnabled(True)



if __name__ == "__main__":
    import sys
    app = QtWidgets.QApplication(sys.argv)
    MainWindow = QtWidgets.QMainWindow()
    ui = Ui_MainWindow()
    ui.setupUi(MainWindow)
    MainWindow.show()
    sys.exit(app.exec_())
